import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox
import pandas as pd
import openpyxl
import os

count = 0

class DataForm:
    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Student Information Form")
        tk.Label(parent, text="Student Name", font="Arial 14 bold").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Department", font="Arial 14 bold").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="DOB", font="Arial 14 bold").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Email ID", font="Arial 14 bold").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Phone No.", font="Arial 14 bold").grid(row=5, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Class 10(%)", font="Arial 14 bold").grid(row=6, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Class 12(%)", font="Arial 14 bold").grid(row=7, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Address", font="Arial 14 bold").grid(row=8, column=0, padx=10, pady=10, sticky="e")
        tk.Label(parent, text="Guardian Name", font="Arial 14 bold").grid(row=9, column=0, padx=10, pady=10, sticky="e")

        self.st_name = ttk.Entry(parent, font="Arial 14")
        self.dept_name = ttk.Entry(parent, font="Arial 14")
        self.emailid = ttk.Entry(parent, font="Arial 14")
        self.phone = ttk.Entry(parent, font="Arial 14")
        self.classX = ttk.Entry(parent, font="Arial 14")
        self.classXII = ttk.Entry(parent, font="Arial 14")
        self.ad = ttk.Entry(parent, font="Arial 14")
        self.g_name = ttk.Entry(parent, font="Arial 14")


        # Adding DOB Combobox
        self.dob_day = ttk.Combobox(parent, values=[str(i) for i in range(1, 32)], font="Arial 14", width=5)
        self.dob_month = ttk.Combobox(parent, values=[str(i) for i in range(1, 13)], font="Arial 14", width=5)
        self.dob_year = ttk.Combobox(parent, values=[str(i) for i in range(1990, 2031)], font="Arial 14", width=5)

        self.st_name.grid(row=1, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.dept_name.grid(row=2, column=1, padx=10, pady=10, ipadx=20, ipady=5)


        self.dob_day.grid(row=3, column=1, padx=5, pady=10, ipady=5, ipadx=94)
        self.dob_month.grid(row=3, column=2, padx=5, pady=10, ipady=5, ipadx=94)
        self.dob_year.grid(row=3, column=3, padx=5, pady=10, ipady=5, ipadx=94)

        self.emailid.grid(row=4, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.phone.grid(row=5, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.classX.grid(row=6, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.classXII.grid(row=7, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.ad.grid(row=8, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        self.g_name.grid(row=9, column=1, padx=10, pady=10, ipadx=20, ipady=5)
        
        back_button = tk.Button(parent, text="Back", command=parent.destroy, font="Arial 14 bold",
                                bg="red", fg="white")
        back_button.grid(row=10, columnspan=3, pady=20)

        submit_button = tk.Button(parent, text="Submit", command=self.submit_form, font="Arial 14 bold",
                                  bg="green", fg="white")
        submit_button.grid(row=10, columnspan=4, pady=20)

        global count
        count = count+1

    def submit_form(self):
        student_name = self.st_name.get()
        department = self.dept_name.get()
        dob = f"{self.dob_day.get()}-{self.dob_month.get()}-{self.dob_year.get()}"
        email = self.emailid.get()
        phone_number = self.phone.get()
        class_10_percentage = self.classX.get()
        class_12_percentage = self.classXII.get()
        address = self.ad.get()
        guardian = self.g_name.get()
        global count
        registration_no = count


        if student_name and department and dob and email and phone_number and class_10_percentage and class_12_percentage and address and guardian:

            data = {
                'Student Name': [student_name],
                'Department': [department],
                'DOB': [dob],
                'Email ID': [email],
                'Phone No.': [phone_number],
                'Class 10(%)': [class_10_percentage],
                'Class 12(%)': [class_12_percentage],
                'Address': [address],
                'Guardian Name': [guardian],
                'Registration No.': [registration_no]
            }

            df = pd.read_excel('Student_Data.xlsx')
            if df.empty:
                pd.DataFrame(data).to_excel('Student_Data.xlsx', index=False)

            else:
                existing_data = pd.read_excel('Student_Data.xlsx')
                c = 0
                workbook = openpyxl.load_workbook('Student_Data.xlsx')
                sheet = workbook.active
                for row in sheet:
                    if not all([cell.value == None for cell in row]):
                        c+=1
                count = sheet.cell(row=c, column=10).value
                data['Registration No.'] = count+1
                updated_data = pd.concat([existing_data, pd.DataFrame(data)], ignore_index=True)
                updated_data.to_excel('Student_Data.xlsx', index=False)
            
            messagebox.showinfo(title="Success", message="Data added successfully !")

        else:
            messagebox.showwarning(title="Error", message="All fields are required")

class MainApp:
    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Main Menu")

        frame = tk.Frame(parent)
        frame.pack(expand=True)

        logo = tk.PhotoImage(file=r"STCET-Logo.png")
        w = tk.Label(frame, image=logo)
        w.image = logo
        w.pack(pady=10)

        topic=tk.Label(frame, text="Student Info System of STCET", font="Arial 20 bold", fg="blue")
        topic.pack(pady=20)

        grp_member=tk.Label(frame, text="Group Members : Mayank Banerjee, Souvagya Dey, Milapan De", font="Arial 18 bold", fg="green")
        grp_member.pack(pady=20)

        options=tk.Frame(frame)
        options.pack(pady=30)

        add_button = tk.Button(options, text="Add", command=self.open_form, font="Arial 14 bold", bg="black", fg="white")
        add_button.pack(side=tk.LEFT, padx=20)

        display_button = tk.Button(options, text="Display", command=self.display_data, font="Arial 14 bold", bg="black", fg="white")
        display_button.pack(side=tk.LEFT, padx=20)

        exit_button = tk.Button(options, text="Exit", command=parent.destroy, font="Arial 14 bold", bg="red", fg="white")
        exit_button.pack(side=tk.LEFT, padx=20)


    def open_form(self):
        form_window = tk.Toplevel(self.parent)
        DataForm(form_window)

    def display_data(self):
        df = pd.read_excel('Student_Data.xlsx')
        if df.empty:
            messagebox.showinfo(title="Alert!", message="No data has been found !")
        else:
            path = "Student_Data.xlsx"
            display_window = tk.Tk()
            display_window.title('Excel Sheet Data')

            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            list_of_values = list(sheet.values)

            cols = list_of_values[0]
            tree = ttk.Treeview(display_window, columns=cols, show="headings")
            # Headings
            for col_name in cols:
                tree.heading(col_name, text=col_name)

            # Displaying the data
            for value_tuple in list_of_values[1:]:
                tree.insert('', tk.END, values=value_tuple)

            # Creating horizontal scrollbar
            horizontal_scroll = ttk.Scrollbar(display_window, orient='horizontal', command=tree.xview)
            horizontal_scroll.pack(side='bottom', fill='x')

            tree.configure(xscrollcommand=horizontal_scroll.set)

            # Creating vertical scrollbar
            vertical_scroll = ttk.Scrollbar(display_window, orient='vertical', command=tree.yview)
            vertical_scroll.pack(side='right', fill='y')

            tree.pack(expand=True, fill='both')

            workbook.close()

if __name__ == "__main__":
    root = tk.Tk()
    MainApp(root)
    root.mainloop()